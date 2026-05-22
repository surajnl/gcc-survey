from flask import Flask, request, jsonify, send_from_directory, send_file
from flask_cors import CORS
import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__, static_folder='public')
CORS(app)

DATA_FILE = 'submissions.json'
ADMIN_PASSWORD = 'admin123'

def load_submissions():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, 'r') as f:
            return json.load(f)
    return []

def save_submissions(data):
    with open(DATA_FILE, 'w') as f:
        json.dump(data, f, indent=2)

@app.route('/')
def index():
    return send_from_directory('public', 'index.html')

@app.route('/api/submit', methods=['POST'])
def submit_survey():
    try:
        data = request.json
        submissions = load_submissions()

        submission = {
            'id': len(submissions) + 1,
            'timestamp': datetime.now().isoformat(),
            'yoe': data.get('yoe', ''),
            'gcc_age': data.get('gcc_age', ''),
            'sector': data.get('sector', ''),
            'rankings': data.get('rankings', [])
        }

        submissions.append(submission)
        save_submissions(submissions)

        return jsonify({'success': True, 'message': 'Survey submitted successfully!'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/download', methods=['POST'])
def download_excel():
    try:
        data = request.json
        if data.get('password') != ADMIN_PASSWORD:
            return jsonify({'success': False, 'error': 'Invalid password'}), 401

        submissions = load_submissions()

        if not submissions:
            return jsonify({'success': False, 'error': 'No submissions yet'}), 404

        wb = Workbook()
        wb.remove(wb.active)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Sheet 1: Profile Data
        ws1 = wb.create_sheet("Profile Data")
        headers1 = ['Submission ID', 'Timestamp', 'Years of Experience', 'GCC Age', 'Sector']
        ws1.append(headers1)

        for cell in ws1[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for sub in submissions:
            ws1.append([
                sub['id'],
                sub['timestamp'],
                sub['yoe'],
                sub['gcc_age'],
                sub['sector']
            ])

        for col in ws1.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws1.column_dimensions[column].width = adjusted_width

        # Sheet 2: All Rankings
        ws2 = wb.create_sheet("All Rankings")
        headers2 = ['Submission ID', 'Timestamp', 'Rank', 'Factor']
        ws2.append(headers2)

        for cell in ws2[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for sub in submissions:
            for rank_item in sub['rankings']:
                ws2.append([
                    sub['id'],
                    sub['timestamp'],
                    rank_item['rank'],
                    rank_item['factor']
                ])

        for col in ws2.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            ws2.column_dimensions[column].width = adjusted_width

        # Sheet 3: Summary
        ws3 = wb.create_sheet("Summary (Avg Rank)")
        headers3 = ['Rank', 'Factor', 'Average Rank', 'Total Votes']
        ws3.append(headers3)

        for cell in ws3[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        factor_scores = {}
        for sub in submissions:
            for rank_item in sub['rankings']:
                factor = rank_item['factor']
                rank = rank_item['rank']
                if factor not in factor_scores:
                    factor_scores[factor] = []
                factor_scores[factor].append(rank)

        summary_data = []
        for factor, scores in factor_scores.items():
            avg_rank = sum(scores) / len(scores)
            summary_data.append({
                'factor': factor,
                'avg_rank': round(avg_rank, 2),
                'votes': len(scores)
            })

        summary_data.sort(key=lambda x: x['avg_rank'])

        for idx, item in enumerate(summary_data, 1):
            ws3.append([
                idx,
                item['factor'],
                item['avg_rank'],
                item['votes']
            ])

        for col in ws3.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            ws3.column_dimensions[column].width = adjusted_width

        filename = 'survey_data.xlsx'
        wb.save(filename)

        return send_file(
            filename, 
            as_attachment=True, 
            download_name=f'GCC_Survey_Data_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# NEW: Clear all data endpoint
@app.route('/api/clear', methods=['POST'])
def clear_data():
    try:
        data = request.json
        if data.get('password') != ADMIN_PASSWORD:
            return jsonify({'success': False, 'error': 'Invalid password'}), 401

        if os.path.exists(DATA_FILE):
            os.remove(DATA_FILE)

        return jsonify({'success': True, 'message': 'All data cleared successfully'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/count', methods=['GET'])
def get_count():
    submissions = load_submissions()
    return jsonify({'count': len(submissions)})

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
